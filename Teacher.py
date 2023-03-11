import leancloud
import PySimpleGUI as sg
from account import log_in
from datetime import datetime
from account import read_key

leancloud.init("OzN2cISaG1cUDK9wLAw2lB4F-gzGzoHsz", "ex4DGUuGw9yQAVoRfwUpbU2p") #通过密钥初始化leancloud

def read_class_key():
    import json
    try:
        with open('data/Class.json','r') as f:
            key=json.loads(f.read())
        Class_ID=key['Class_ID']
        Class_Name=key['Class_Name']
        return Class_ID,Class_Name
    except:
        return False

def Get_Class_ID() -> str:
    '''
    获取到可用的班级ID
    '''
    import random
    Class_ID=random.randint(1000,9999)
    print('Class_ID -> ',Class_ID)
    Class_ID='C'+str(Class_ID)
    Class_list_services = leancloud.Object.extend('Class_list') #定位到Class_list类
    Class_list_service = Class_list_services.query #查询服务
    Class_ID_Status = Class_list_service.equal_to('Class_ID',Class_ID).find() #查询该班级ID是否被占用
    if Class_ID_Status == []:
        return Class_ID
    else:
        Get_Class_ID()

def Get_Plan_ID() -> str:
    '''
    获取到可用的任务ID
    '''
    import random
    HomeworkID=read_class_key()[0]+'-'+str(random.randint(10000,99999))
    print('HomeworkID -> ',HomeworkID)
    Homework_list_services = leancloud.Object.extend('Homework_list') #定位到Class_list类
    Homework_list_services = Homework_list_services.query #查询服务
    Homework_ID_Status = Homework_list_services.equal_to('HomeworkID',HomeworkID).find() #查询该任务ID是否被占用
    if Homework_ID_Status == []:
        return HomeworkID
    else:
        Get_Plan_ID()

class Creat_Class():
    '''
    创建班级
    '''
    def __init__(self) -> None:
        pass

    def Creat_student(self,Realname:str,Class_ID:str) -> None:
        '''
        注册学生
        '''
        from account import sign_up
        import pypinyin
        En_Username = ''
        for i in pypinyin.pinyin(Realname, style=pypinyin.NORMAL):
            En_Username += ''.join(i)
        Username = Class_ID+En_Username
        sign_up(User_name=Username,password=Username)
        UserID = log_in(User_name=Username,password=Username)
        User_services = leancloud.Object.extend('Class_member')
        User_service=User_services()
        User_service.set('Class_ID',Class_ID)
        User_service.set('User_ID',UserID)
        User_service.set('Realname',Realname)
        User_service.save()
        return Username

    def save_key(self,Class_ID=None,Class_Name=None):
        import json
        key={
            'Class_ID':Class_ID,
            'Class_Name':Class_Name
        }
        with open(r'data\Class.json','w') as f:
            f.write(json.dumps(key))
        return None

    def Layout_GUI(self):
        import os,subprocess
        path=os.getenv('APPDATA')+'\Study to do(Teacher)'
        file_path=(path+'\data\Class.xlsx')
        print(file_path)
        layout=[
            [sg.Text('学生账号和密码已保存, 请到{}查看'.format(file_path))],
            [sg.Button('打开'),sg.Button('打开文件夹'),sg.Push(),sg.Button('退出')]
        ]
        Done_GUI=sg.Window('导出完成',layout=layout,icon='ico/LOGO.ico')
        event=Done_GUI.Read()
        Done_GUI.Close()
        if event[0]=='打开':
            from tool import open_file_tip
            Twindow = open_file_tip()      
            subprocess.getstatusoutput('"{}"'.format(file_path))
            Twindow.Close()        
        elif event[0]=='打开文件夹':
            file_path=os.path.dirname(file_path)
            subprocess.getstatusoutput('explorer.exe '+file_path)
        return None

    def GUI(self) -> None:
        '''
        GUI支持服务
        '''
        if self.CheckIfBindingAClass():
            return None
        from openpyxl import load_workbook,Workbook
        from tool import progress
        import os
        layout = [
            [sg.Text('创建班级',font=('微软雅黑 15'))],
            [sg.Text('注意:在早期开发版本中,请将您所属班级全部添加至一个相同班级中',text_color='red')],
            [sg.Text('班级名: '),sg.Push(),sg.Input()],
            [sg.Text('从Excel(.xlsx)文件中导入学生姓名: '),sg.Input(),sg.FileBrowse('打开',file_types=(("xlsx 文件", "*.xlsx"),))],
            [sg.Text('文档格式规范要求和模板文档下载: '),sg.Button('跳转')],
            [sg.VPush()],
            [sg.Push(),sg.Button('下一步')]
        ]
        window = sg.Window('新建班级',layout=layout,icon='ico/LOGO.ico',font=('微软雅黑 10'),size=(650,225))
        event, value = window.Read()
        if event == sg.WIN_CLOSED:
            window.Close()
            return None
        if event == '跳转':
            window.Close()
            import webbrowser
            webbrowser.open_new('https://doc.muspace.top/#/zh-cn/Program/Study-To-Do?id=创建班级')
            return None
        self.Class_Name = value[0]
        Excel_File = value[1]
        print(Excel_File)
        if self.Class_Name=='' or Excel_File=='':
            sg.Popup('所有的输入框不能为空!')
            window.Close()
            return None
        if os.path.splitext(Excel_File)[-1] != '.xlsx':
            sg.Popup('文档不符合导入格式要求, 请使用.xlsx文档导入!')
            window.Close()
            return None
        print(value)
        sheet = load_workbook(Excel_File).worksheets[0]
        col=[]
        for a in sheet['A']:
            col.append(a.value)
        del col[0]
        if len(col) == 0:
            sg.Popup('学生数为0,请检查')
            window.Close()
            return None
        p = progress()
        p.new()
        self.Class_ID = Get_Class_ID()
        p.add(25)
        self.main()
        p.add(50)
        #导出Excel文件
        wb = Workbook()    #创建文件对象 
        #获取第一个sheet
        ws = wb.active 
        # 将数据写入到指定的单元格
        ws['A1'] = '姓名'
        ws['B1'] = '账号'
        ws['C1'] = '密码'
        i=1
        for stu in col:
            if stu == None or stu == '':
                continue
            i+=1
            Username = self.Creat_student(stu,self.Class_ID)
            print(Username)
            ws['A'+str(i)]=stu
            ws['B'+str(i)]=Username
            ws['C'+str(i)]=Username
        wb.save("data/Class.xlsx")
        p.add(100)
        p.close()
        self.save_key(Class_ID=self.Class_ID,Class_Name=self.Class_Name)
        self.Layout_GUI()
        window.Close()

    def main(self):
        '''
        提交数据至远程服务端
        '''
        from datetime import datetime
        from account import read_key,log_in
        User_name, Password = read_key()
        User_ID = log_in(User_name,Password)
        Creat_Time = datetime.strftime(datetime.now(),'%Y-%m-%d %H:%M')
        Class_list_services = leancloud.Object.extend('Class_list') #定位到Class_list类
        Class_list_service = Class_list_services() #初始化类
        Class_list_service.set('Class_ID',self.Class_ID)
        Class_list_service.set('Class_Name',self.Class_Name)
        Class_list_service.set('Creat_Time',Creat_Time)
        Class_list_service.set('Teacher',User_ID)
        Class_list_service.save()

    def CheckIfBindingAClass(self):
        '''
        检查是否已经绑定一个班级
        '''
        from account import read_key,log_in
        from plan_manager import save_plan
        User_name, Password = read_key()
        User_ID = log_in(User_name,Password)
        Class_list_services = leancloud.Object.extend('Class_list').query
        Class_list_services.equal_to('Teacher',User_ID)
        Class_list = Class_list_services.find()
        if len(Class_list) == 0:
            return False
        else:
            from tool import progress
            p = progress()
            p.new()
            Class_id = Class_list[0].get('Class_ID')
            Class_name = Class_list[0].get('Class_Name')
            self.save_key(Class_ID=Class_id,Class_Name=Class_name)
            #查找该班级下的作业
            Plan_list = leancloud.Object.extend('Homework_list').query.equal_to('Class_ID',Class_id).find()
            for index,plan in enumerate(Plan_list):
                p.add(int(index/len(Plan_list)*100))
                HomeworkID = plan.get('HomeworkID')
                des = plan.get('des')
                etime = plan.get('etime')
                stime = plan.get('stime')
                sub = plan.get('sub')
                title = plan.get('plan')
                textname = HomeworkID + '.json'
                save_plan(textname,sub,title,stime,etime,des,status='进行中',HomeworkID=HomeworkID)
            p.close()
            sg.Popup('已经绑定班级并且成功同步数据')
            return True

def View_Class():
    '''
    查看班级信息
    '''
    Class_ID, Class_Name = read_class_key()
    # 获取班级成员
    Class_services = leancloud.Object.extend('Class_member').query
    Class_services.equal_to('Class_ID', Class_ID)
    Class_list = Class_services.find()
    stu=[]
    for plan in Class_list:
        stu.append(plan.get('Realname'))
    # 班级成员格式化
    text = ', '
    texts = ''
    while stu!=[]:
        print('stu ->',stu)
        texts+=(text.join(stu[:10])+'\n')
        del stu[:10]
    layout = [
        [sg.Text('班级名:',font=('微软雅黑 15')),sg.Text(Class_Name)],
        [sg.Text('班级ID:',font=('微软雅黑 15')),sg.Text(Class_ID)],
        [sg.Text('班级成员:',font=('微软雅黑 15'))],
        [sg.Text(texts)],
        [sg.Push(),sg.Button('退出')]
    ]
    window = sg.Window('班级信息',layout=layout,font=('微软雅黑 10'))
    window.Read()
    window.Close()
    return None

def Send_Plan(HomeworkID=str) -> None:
    '''
    发布/编辑/完成任务
    '''
    from plan_manager import get_plan_info
    import os
    Class_ID = read_class_key()[0]
    HomeworkID,des,etime,status,stime,sub,plan=get_plan_info(os.listdir('plan').index(HomeworkID+'.json'),'all')
    Homework_list_services = leancloud.Object.extend('Homework_list') #定位到Homework_list类
    Homework_list_service = Homework_list_services() #初始化类
    Homework_list_service.set('Class_ID',Class_ID)
    Homework_list_service.set('HomeworkID',HomeworkID)
    Homework_list_service.set('sub',sub)
    Homework_list_service.set('plan',plan)
    Homework_list_service.set('stime',stime)
    Homework_list_service.set('etime',etime)
    Homework_list_service.set('des',des)
    Homework_list_service.save()

def Get_OtFinishStus(HomeworkID:str) -> str:
    '''
    获取超时完成学生列表
    :HomeworkID 作业ID
    返回字符串
    '''
    stus = []
    #获取完成学生列表
    Homework_Status_services = leancloud.Object.extend('Homework_Status').query
    Homework_Status_services.equal_to('HomeworkID', HomeworkID)
    Homework_Status_services.equal_to('Overtime',True)
    Homework_Status_list = Homework_Status_services.find()
    #移除已完成学生
    for h in Homework_Status_list:
        print('Remove -> ',h.get('Realname'))
        stus.append(h.get('Realname'))
    #班级成员格式化
    text = ', '
    texts = ''
    while stus!=[]:
        texts+=(text.join(stus[:10])+'\n')
        del stus[:10]
    if texts=='':
        texts='空'
    return texts
  
def Get_UnFinishStus(HomeworkID:str) -> str:
    '''
    查看任务信息,但是返回未完成字符串
    '''
    #获取班级成员
    Class_services = leancloud.Object.extend('Class_member').query
    Class_services.equal_to('Class_ID', read_class_key()[0])
    Class_list = Class_services.find()
    stus=[]
    for plan in Class_list:
        stus.append(plan.get('Realname'))
    print('stus -> ',stus)
    #获取完成学生列表
    Homework_Status_services = leancloud.Object.extend('Homework_Status').query
    Homework_Status_services.equal_to('HomeworkID', HomeworkID)
    Homework_Status_list = Homework_Status_services.find()
    #移除已完成学生
    for h in Homework_Status_list:
        print('Remove -> ',h.get('Realname'))
        stus.remove(h.get('Realname'))
    #班级成员格式化
    text = ', '
    texts = ''
    while stus!=[]:
        texts+=(text.join(stus[:10])+'\n')
        del stus[:10]
    if texts=='':
        texts='(所有学生均已完成)'
    return texts

class Class_Chat:
    def Time_Get(datetimes=False):
        '''
        返回时间字符串
        '''
        if not datetimes:
            datetimes=datetime.now()
        time_str = datetime.strftime(datetimes,'%m-%d %H:%M:%S')
        return time_str

    def Format_Message(CreateAt,Username,Message):
        '''
        格式化信息
        '''
        return CreateAt+' '+Username+' 说: '+Message

    def Format_Message_From_Datetime(CreateAt,Username,Message):
        '''
        格式化信息
        '''
        CreateAt = datetime.strftime(CreateAt,'%m-%d %H:%M:%S')
        return CreateAt+' '+Username+' 说: '+Message
    
    def Get_New_Five_Message(ClassID):
        '''
        从云端获取信息后格式化
        :Return: 格式化后的5条新信息和一条最新信息
        '''
        Messages_list,Usernames_list,CreatAt_list = Class_Chat.get_message(ClassID)
        NewMessage_str = ''
        for num in range(len(Messages_list)):
            NewMessage_str += Class_Chat.Format_Message(Class_Chat.Time_Get(CreatAt_list[num]),Usernames_list[num],Messages_list[num]) + '\n'
        if len(Messages_list) >= 1:
            LastestMessage_str = Class_Chat.Format_Message(Class_Chat.Time_Get(CreatAt_list[-1]),Usernames_list[-1],Messages_list[-1])
        else:
            LastestMessage_str = ''
        return NewMessage_str,LastestMessage_str

    # 获取云端信息
    def get_message(ClassID) -> list:
        '''
        获取到最新的前五条信息
        :param ClassID: 班级ID
        :return 返回最新的前五条信息,顺序为Message,Username,CreatedAt
        '''
        # 绑定到MailBox类
        class_obj = leancloud.Object.extend('MailBox')
        # 查询服务
        class_obj_query = class_obj.query
        # 查询该班级ID的信息
        class_obj_query.equal_to('Class_ID',ClassID)
        # 获取信息
        class_obj_query_result = class_obj_query.find()
        # 返回信息
        # 检查是否有信息
        if class_obj_query_result == []:
            return [],[],[]
        else:
            Messages_list,Usernames_list,CreatAt_list = [],[],[]
            for i in class_obj_query_result:
                Messages_list.append(i.get('Message'))
                Usernames_list.append(i.get('Username'))
                CreatAt_list.append(i.get('createdAt'))
            return Messages_list,Usernames_list,CreatAt_list

    #发送信息至云端
    def send_message(message):
        '''
        返回一个时间字符串以校准时间
        '''
        ClassID,_ = read_class_key()
        Username = read_key()[0]
        # 绑定到MailBox类
        class_obj = leancloud.Object.extend('MailBox')
        class_obj = class_obj()
        # 发送信息,包括班级ID和用户名
        class_obj.set('Message',message)
        class_obj.set('Class_ID',ClassID)
        class_obj.set('Username',Username)
        class_obj.save()
        return datetime.strftime(class_obj.get('createdAt'),'%m-%d %H:%M:%S')

    # 实时获取云端信息
    def get_realtime_message(ClassID,OldMessage) -> str:
        '''
        获取到最新的一条信息
        :param ClassID: 班级ID
        :return 返回最新的一条信息,顺序为Message,Username,createdAt
        '''
        # 绑定到MailBox类
        class_obj = leancloud.Object.extend('MailBox')
        # 查询服务
        class_obj_query = class_obj.query
        # 查询该班级ID的信息
        class_obj_query.equal_to('Class_ID',ClassID)
        # 获取信息
        class_obj_query_result = class_obj_query.find()
        if class_obj_query_result == []:
            return ''
        else:
            Formated_Message = Class_Chat.Format_Message_From_Datetime(class_obj_query_result[-1].get('createdAt'),class_obj_query_result[-1].get('Username'),class_obj_query_result[-1].get('Message'))
            if Formated_Message != OldMessage:
                return Formated_Message
            else:
                return ''

    def main():
        ClassID,_ = read_class_key()
        Username = read_key()[0]
        Class_list_services = leancloud.Object.extend('Class_list') #定位到Class_list类
        Class_list_service = Class_list_services.query #查询服务
        Class_ID_Status = Class_list_service.equal_to('Class_ID',ClassID).find()
        if len(Class_ID_Status) == 0:
            sg.Popup('该账号尚未绑定任何班级!',font=('微软雅黑 10'))
            return None
        NewMessage_str,Old_Message = Class_Chat.Get_New_Five_Message(ClassID)
        layout =[
            [sg.Text('消息框')],
            [sg.Multiline(size=(80,20),key='_OUTPUT_',default_text=NewMessage_str,disabled=True,autoscroll=True,text_color='blue')],
            [sg.Text('发送新消息:')],
            [sg.Input(size=(75,5),key='_INPUT_'),sg.Button('发送')],
        ]
        window = sg.Window('消息框',layout,font=('微软雅黑 10'),size=(700,500),return_keyboard_events=True)
        while True:
            event,value = window.Read(timeout=1500)
            if event == sg.WIN_CLOSED:
                window.Close()
                return None
            elif event in ('\r','发送'):
                Message = value['_INPUT_']
                Message_Time = Class_Chat.send_message(Message)
                Old_Message = Class_Chat.Format_Message(Message_Time,Username,Message)
                window.Element('_OUTPUT_').Update(Old_Message+'\n',text_color_for_value='black',append=True)
                window.Element('_INPUT_').Update('')
            elif event == '__TIMEOUT__':
                print('Updating Message...')
                Realtime_message = Class_Chat.get_realtime_message(ClassID,Old_Message)
                if Realtime_message != '':
                    Old_Message = Realtime_message
                    window.Element('_OUTPUT_').Update(Realtime_message+'\n',text_color_for_value='red',append=True)