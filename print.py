import pdfkit,os,subprocess
from markdown import markdown
import PySimpleGUI as sg
from plan_manager import *

def wkhtmltox_check():
    try:
        os.chdir('wkhtmltox')
        os.chdir('..')
        return True
    except:
        layout=[
            [sg.Text('该功能需要第三方应用程序: wkhtmltox 的支持',font=('微软雅黑 10'))],
            [sg.Text('请问要现在下载吗? ( 46M 预计1分钟以内 )',font=('微软雅黑 10'))],
            [sg.Button('是',font=('微软雅黑 10')),sg.Button('否',font=('微软雅黑 10'))]
        ]
        windows=sg.Window('第三方软件下载提示',layout=layout,icon='ico/LOGO.ico')
        event,value=windows.Read()
        if event==sg.WIN_CLOSED:
            windows.Close()
            return False
        if event=='是':
            windows=sg.Window('下载中',layout=[[sg.Text('下载中, 这通常需要一点时间...',font=('微软雅黑 10'))]],icon='ico/LOGO.ico')
            windows.read(timeout=100)
            import requests as r
            import zipfile
            try:
                url='https://file.muspace.top/Program/wkhtmltox.zip'
                file=r.get(url)
                open('wkhtmltox.zip','wb').write(file.content)
                with zipfile.ZipFile('wkhtmltox.zip') as zf:
                    zf.extractall()
                os.remove('wkhtmltox.zip')
                windows.Close()
                sg.Popup('wkhtmltox 下载完毕, 现在请重新点击导出按钮进行导出',font=('微软雅黑 10'))
            except:
                sg.Popup('wkhtmltox 下载失败, 请检查是否已联网或者是已关闭代理',font=('微软雅黑 10'))
            return True
        else:
            windows.Close()
            return False

def markdown_to_pdf(file_path):
    input_filename = file_path.replace('pdf','md')
    output_filename = file_path
    with open(input_filename, encoding='utf-8') as f:
        text = f.read()
    html = markdown(text, output_format='html')  # MarkDown转HTML
    wkhtmltopdf = r'wkhtmltox\bin\wkhtmltopdf.exe'  # 指定wkhtmltopdf
    configuration = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf)
    pdfkit.from_string(html, output_filename, configuration=configuration, options={'encoding': 'utf-8'})  # HTML转PDF

def txt_to_markdown(sub,File_path):
    text_name=read_plan_with_setting(sub)
    text=''
    if text_name==[]:
        sg.Popup('您的导出范围有误, 里面没有任何任务!',font=('微软雅黑 10'))
        return 'Error'
    if sub=='全部':
        text+='# 所有任务'
    else:
        text+='# 所有任务({})'.format(sub.split('-')[1])
    text+='\n'
    num=0
    a=0
    for num in text_name:
        status=get_plan_info(num,'status')
        if status!='已结束':
            a+=1
            text+='# □'+str(a)+'. '+get_plan_info(num,'sub')+': '+get_plan_info(num,'text')+'\n'
            atime=get_plan_info(num,'atime')
            if atime!=(None,None):
                text+='> 从 '+atime[0]+' 至 '+atime[1]+'\n'
    if a==0:
        sg.Popup('您的导出范围有误, 里面没有任何任务!',font=('微软雅黑 10'))
        return 'Error'
    with open(File_path,'w',encoding='utf-8') as f:
        f.write(text)
    return None

def txt_to_markdown_GUI():
    wkhtmltox_check()
    layout=[
        [sg.Text('选择您所导出的任务范围:',font=('微软雅黑 10')),sg.InputCombo(['全部','按科目-语文','按科目-数学','按科目-英语','按科目-物理','按科目-化学','按科目-生物','按科目-地理','按科目-历史','按科目-政治','按科目-体育','按分类-生活','按时间-当天','按时间-当周'],font=('微软雅黑 10'),size=(10,5))],
        [sg.Text('导出到:',font=('微软雅黑 10')),sg.Input(),sg.FileSaveAs('浏览',font=('微软雅黑 10'),file_types=(("Markdown 文件", "*.md"), ))],
        [sg.Button('提交',font=('微软雅黑 10'))]
    ]
    windows=sg.Window('导出为Markdown文件',layout=layout,icon='ico/LOGO.ico')
    event,value=windows.Read()
    windows.Close()
    if event==sg.WIN_CLOSED:
        return None
    sub=value[0]
    File_path=value[1]
    if sub=='' or File_path=='':
        sg.Popup('任务范围/路径不可为空!')
        return None
    if txt_to_markdown(sub,File_path)=='Error':
        return None
    layout=[
        [sg.Text('导出完成,已导出到 '+File_path+' 中',font=('微软雅黑 10'))],
        [sg.Text('请选择接下来的操作:',font=('微软雅黑 10'))],
        [sg.Button('打开文件',font=('微软雅黑 10')),sg.Button('打开所在文件夹'),sg.Push(),sg.Button(tooltip='返回',button_color=(sg.theme_background_color(), sg.theme_background_color()),border_width=0,image_filename='ico/back.png',key='返回')]
    ]
    done_window=sg.Window('导出完成',layout=layout,icon='ico/LOGO.ico',font=('微软雅黑 10'))
    event=done_window.Read()
    print(File_path)
    if event[0]=='打开文件':
        from tool import open_file_tip
        Twindow = open_file_tip()      
        subprocess.getstatusoutput(File_path)
        Twindow.Close()
    elif event[0]=='打开所在文件夹':
        subprocess.getstatusoutput('explorer.exe '+os.path.dirname(File_path))
    else:
        pass
    done_window.Close()
    return None

def txt_to_PDF_GUI():
    if not wkhtmltox_check():
        return None
    layout=[
        [sg.Text('选择您所导出的任务范围:',font=('微软雅黑 10')),sg.InputCombo(['全部','按科目-语文','按科目-数学','按科目-英语','按科目-物理','按科目-化学','按科目-生物','按科目-地理','按科目-历史','按科目-政治','按科目-体育','按分类-生活','按时间-当天','按时间-当周','按状态-已超时'],font=('微软雅黑 10'),size=(10,5))],
        [sg.Text('导出到:',font=('微软雅黑 10')),sg.Input(),sg.FileSaveAs('浏览',font=('微软雅黑 10'),file_types=(("PDF 文件", "*.pdf"), ))],
        [sg.Button('提交',font=('微软雅黑 10'))]
    ]
    windows=sg.Window('导出为Pdf文件',layout=layout,icon='ico/LOGO.ico',font=('微软雅黑 10'))
    event,value=windows.Read()
    windows.Close()
    if event==sg.WIN_CLOSED:
        return None
    sub=value[0]
    File_path=value[1]
    if sub=='' or File_path=='':
        sg.Popup('任务范围/路径不可为空!')
        return None
    if txt_to_markdown(sub,File_path.replace('pdf','md'))=='Error':
        return None
    markdown_to_pdf(File_path)
    os.remove(File_path.replace('pdf','md'))
    layout=[
        [sg.Text('导出完成,已导出到 '+File_path+' 中')],
        [sg.Text('请选择接下来的操作:')],
        [sg.Button('打开文件'),sg.Button('打开所在文件夹'),sg.Push(),sg.Button('返回')]
    ]
    done_window=sg.Window('导出完成',layout=layout,icon='ico/LOGO.ico')
    event=done_window.Read()
    if event[0]=='打开文件':
        from tool import open_file_tip
        Twindow = open_file_tip()      
        subprocess.getstatusoutput(File_path)
        Twindow.Close()
    elif event[0]=='打开所在文件夹':
        subprocess.getstatusoutput('explorer.exe '+os.path.dirname(File_path))
    else:
        pass
    done_window.Close()
    return None

class plan_log:
    def check_value(value):
        if value==None:
            return '无'
        else:
            return value

    def repair_value(value):
        for num in range(len(value)):
            value[num]=(plan_log.check_value(value[num]))
        return value

    def write_column(value,abc,ws):
        for i in range(len(value)):
            ws[abc+str(i+2)]=value[i]
        return ws
    
    def GUI():
        import sys,os
        path=os.getenv('APPDATA')+'\Study to do'
        file_path=(path+'\data\Plan_log.xlsx')
        print(file_path)
        layout=[
            [sg.Text('导出成功, 请到{}查看'.format(file_path))],
            [sg.Button('打开'),sg.Button('打开文件夹'),sg.Push(),sg.Button(tooltip='返回',button_color=(sg.theme_background_color(), sg.theme_background_color()),border_width=0,image_filename='ico/back.png',key='返回')]
        ]
        Done_GUI=sg.Window('导出完成',layout=layout,icon='ico/LOGO.ico')
        event=Done_GUI.Read()
        Done_GUI.Close()
        if event[0]=='打开':
            from tool import open_file_tip
            Twindow = open_file_tip()      
            subprocess.getstatusoutput('"{}"'.format(file_path))
            Twindow.Close()        
        elif event[0]=='打开文件夹':
            file_path=os.path.dirname(file_path)
            subprocess.getstatusoutput('explorer.exe '+file_path)
        return None

    def main(GUI=True):
        from openpyxl import Workbook
        from tool import repair_num
        wb = Workbook()    #创建文件对象 
        #获取第一个sheet
        ws = wb.active     
        # 将数据写入到指定的单元格
        ws['A1'] = '序号'
        ws['B1'] = '科目'
        ws['C1'] = '任务'
        ws['D1'] = '任务开始时间'
        ws['E1'] = '任务结束时间'
        ws['F1'] = '任务描述'
        ws['G1'] = '任务状态'
        txtname_list=read_plan_list()
        plan_num=[]
        plan_sub=[]
        plan_text=[]
        plan_stime=[]
        plan_etime=[]
        plan_des=[]
        plan_status=[]
        for plan_name in range(len(txtname_list)):
            plan_num.append(repair_num(plan_name+1))
            plan_sub.append(get_plan_info(plan_name,'sub'))
            plan_text.append(get_plan_info(plan_name,'text'))
            st,et=get_plan_info(plan_name,'atime')
            plan_stime.append(st)
            plan_etime.append(et)
            plan_des.append(get_plan_info(plan_name,'des'))
            plan_status.append(get_plan_info(plan_name,'status'))
        plan_stime=plan_log.repair_value(plan_stime)
        plan_etime=plan_log.repair_value(plan_etime)
        plan_des=plan_log.repair_value(plan_des)
        ws=plan_log.write_column(plan_num,'A',ws)
        ws=plan_log.write_column(plan_sub,'B',ws)
        ws=plan_log.write_column(plan_text,'C',ws)
        ws=plan_log.write_column(plan_stime,'D',ws)
        ws=plan_log.write_column(plan_etime,'E',ws)
        ws=plan_log.write_column(plan_des,'F',ws)
        ws=plan_log.write_column(plan_status,'G',ws)
        wb.save("data/Plan_log.xlsx")
        if GUI==True:
            plan_log.GUI()
        return None
